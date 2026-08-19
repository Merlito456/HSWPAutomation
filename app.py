import streamlit as st
import pandas as pd
from datetime import datetime
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Set page configuration
st.set_page_config(
    page_title="HSWP Automation Tool",
    page_icon="🛡️",
    layout="wide"
)

st.title("🛡️ Health and Safety Work Permit Automation")
st.markdown("---")

def create_checkbox(value):
    """Return checkbox character"""
    return "☑" if value else "☐"

def create_excel_file(data):
    """Create a clean Excel file from scratch"""
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Permit"
        
        # Define styles
        header_font = Font(name='Arial', size=14, bold=True)
        title_font = Font(name='Arial', size=12, bold=True)
        label_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 25
        
        row = 1
        
        # ============================================
        # HEADER
        # ============================================
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="HEALTH and SAFETY WORK PERMIT")
        cell.font = header_font
        cell.alignment = center_alignment
        row += 1
        
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="NOKIA SHANGHAI BELL")
        cell.font = header_font
        cell.alignment = center_alignment
        row += 1
        
        # ============================================
        # PROJECT DETAILS
        # ============================================
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value="PROJECT DETAILS")
        cell.font = title_font
        cell.alignment = center_alignment
        row += 1
        
        # Row: Sub Contractor
        ws.cell(row=row, column=1, value="Name of Sub Contractor").font = label_font
        ws.cell(row=row, column=2, value=data.get('sub_contractor', 'Ultegra Supplies and Services'))
        ws.cell(row=row, column=4, value="Requesting Vendor").font = label_font
        ws.cell(row=row, column=5, value=data.get('requesting_vendor', 'NOKIA SHANGHAI BELL'))
        row += 1
        
        # Row: Project In-Charge
        ws.cell(row=row, column=1, value="Project In-Charge").font = label_font
        ws.cell(row=row, column=2, value=data.get('project_in_charge', ''))
        ws.cell(row=row, column=4, value="Person In-charge").font = label_font
        ws.cell(row=row, column=5, value=data.get('person_in_charge', 'John Carlo Rabanes'))
        row += 1
        
        # Row: Safety Officer
        ws.cell(row=row, column=1, value="Project Safety Officer").font = label_font
        ws.cell(row=row, column=2, value=data.get('safety_officer', 'RONNIE ALVIN CHIU'))
        ws.cell(row=row, column=4, value="Work Schedule").font = label_font
        ws.cell(row=row, column=5, value=data.get('work_schedule', ''))
        ws.cell(row=row, column=6, value="Work Time Period").font = label_font
        ws.cell(row=row, column=7, value=data.get('work_time_period', ''))
        row += 1
        
        # Row: Project Name / Start Date / Start Time
        ws.cell(row=row, column=1, value="Project Name").font = label_font
        ws.cell(row=row, column=2, value=data.get('project_name', 'FTTH HORIZONTAL'))
        ws.cell(row=row, column=4, value="Start Date").font = label_font
        ws.cell(row=row, column=5, value=data.get('start_date', '08/10/2026'))
        ws.cell(row=row, column=6, value="Start Time").font = label_font
        ws.cell(row=row, column=7, value=data.get('start_time', '08:00AM'))
        row += 1
        
        # Row: Work Location / End Date / End Time
        ws.cell(row=row, column=1, value="Work Location").font = label_font
        ws.cell(row=row, column=2, value=data.get('work_location', 'MIN624_TS ORAPOBBUTUANAGN'))
        ws.cell(row=row, column=4, value="End Date").font = label_font
        ws.cell(row=row, column=5, value=data.get('end_date', '09/10/2026'))
        ws.cell(row=row, column=6, value="End Time").font = label_font
        ws.cell(row=row, column=7, value=data.get('end_time', '08:00PM'))
        row += 1
        
        # Row: Tower Type / Brief Description
        ws.cell(row=row, column=1, value="Tower Type").font = label_font
        ws.cell(row=row, column=2, value=data.get('tower_type', 'ground base'))
        ws.cell(row=row, column=4, value="Brief Description of Work").font = label_font
        ws.merge_cells(f'E{row}:G{row}')
        ws.cell(row=row, column=5, value=data.get('brief_description', 'SFP link upgrade, Survey'))
        row += 2
        
        # ============================================
        # WORK DETAILS
        # ============================================
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="WORK DETAILS")
        cell.font = title_font
        cell.alignment = center_alignment
        row += 1
        
        # High Risk
        high_risk = data.get('high_risk', 'NO')
        ws.cell(row=row, column=1, value="Is work to be done with high risk?").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(high_risk == 'YES'))
        ws.cell(row=row, column=3, value="YES")
        ws.cell(row=row, column=4, value=create_checkbox(high_risk == 'NO'))
        ws.cell(row=row, column=5, value="NO")
        row += 1
        
        # Work at Heights
        ws.cell(row=row, column=1, value="Work at Heights").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(data.get('work_at_heights', False)))
        
        if data.get('work_at_heights', False):
            ws.cell(row=row, column=3, value="scaffold")
            ws.cell(row=row, column=4, value=create_checkbox(data.get('scaffold', False)))
            ws.cell(row=row, column=5, value="ladder")
            ws.cell(row=row, column=6, value=create_checkbox(data.get('ladder', False)))
            ws.cell(row=row, column=7, value="tower")
            ws.cell(row=row, column=8, value=create_checkbox(data.get('tower', False)))
        row += 1
        
        # Certifications
        ws.cell(row=row, column=1, value="NCII Cert of Scaffold Erector").font = label_font
        ws.cell(row=row, column=2, value=data.get('scaffold_cert', ''))
        ws.cell(row=row, column=4, value="WAH Rigger Certificate").font = label_font
        ws.cell(row=row, column=5, value=data.get('wah_rigger_cert', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="Scaffold components available").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(data.get('scaffold_components', False)))
        ws.cell(row=row, column=4, value="Workers physically fit").font = label_font
        ws.cell(row=row, column=5, value=create_checkbox(data.get('workers_fit', False)))
        row += 1
        
        # Electrical Works
        ws.cell(row=row, column=1, value="Electrical Works").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(data.get('electrical_works', False)))
        row += 1
        
        ws.cell(row=row, column=1, value="NCII Cert of Electrician or ID of REE/RME").font = label_font
        ws.cell(row=row, column=2, value=data.get('electrician_cert', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="LOTO Device").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(data.get('loto_device', False)))
        ws.cell(row=row, column=3, value="Insulated Tools").font = label_font
        ws.cell(row=row, column=4, value=create_checkbox(data.get('insulated_tools', False)))
        row += 1
        
        # Heavy Lifting
        ws.cell(row=row, column=1, value="Heavy Lifting w/ Equipment /Excavation Works").font = label_font
        row += 1
        
        ws.cell(row=row, column=1, value="NCII Cert of Operator").font = label_font
        ws.cell(row=row, column=2, value=data.get('operator_cert', ''))
        ws.cell(row=row, column=3, value="Cert of Lift Rigger").font = label_font
        ws.cell(row=row, column=4, value=data.get('rigger_cert', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="3rd Party Certification of Heavy Eqpt").font = label_font
        ws.cell(row=row, column=2, value=data.get('heavy_eqpt_cert', ''))
        row += 1
        
        # Confined Space
        ws.cell(row=row, column=1, value="Confined Space Works").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(data.get('confined_space', False)))
        row += 1
        
        ws.cell(row=row, column=1, value="Certificate of SCBA Operator").font = label_font
        ws.cell(row=row, column=2, value=data.get('scba_cert', ''))
        ws.cell(row=row, column=3, value="Ventilation Equipment").font = label_font
        ws.cell(row=row, column=4, value=data.get('ventilation_eqpt', ''))
        row += 1
        
        ws.cell(row=row, column=1, value="OxyFuel flash back arrester installed").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(data.get('flash_arrester', False)))
        ws.cell(row=row, column=3, value="Fire Blanket").font = label_font
        ws.cell(row=row, column=4, value=create_checkbox(data.get('fire_blanket', False)))
        row += 1
        
        ws.cell(row=row, column=1, value="O2 and Gas Detector").font = label_font
        ws.cell(row=row, column=2, value=data.get('o2_detector', ''))
        ws.cell(row=row, column=3, value="Safety Line").font = label_font
        ws.cell(row=row, column=4, value=data.get('safety_line', ''))
        row += 1
        
        # Harmful Substances
        harmful = data.get('harmful_substance', 'NO')
        ws.cell(row=row, column=1, value="Is there any harmful substance or nuisance release?").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(harmful == 'YES'))
        ws.cell(row=row, column=3, value="YES")
        ws.cell(row=row, column=4, value=create_checkbox(harmful == 'NO'))
        ws.cell(row=row, column=5, value="NO")
        row += 1
        
        if harmful == 'YES':
            ws.cell(row=row, column=1, value="Fumes").font = label_font
            ws.cell(row=row, column=2, value=create_checkbox(data.get('fumes', False)))
            ws.cell(row=row, column=3, value="Offensive Odors").font = label_font
            ws.cell(row=row, column=4, value=create_checkbox(data.get('odors', False)))
            row += 1
            
            ws.cell(row=row, column=1, value="Dust").font = label_font
            ws.cell(row=row, column=2, value=create_checkbox(data.get('dust', False)))
            ws.cell(row=row, column=3, value="Noise").font = label_font
            ws.cell(row=row, column=4, value=create_checkbox(data.get('noise', False)))
            row += 1
            
            ws.cell(row=row, column=1, value="Sparks").font = label_font
            ws.cell(row=row, column=2, value=create_checkbox(data.get('sparks', False)))
            ws.cell(row=row, column=3, value="Others:").font = label_font
            ws.cell(row=row, column=4, value=data.get('other_harmful', ''))
            row += 1
        
        # Utility Interruption
        utility = data.get('utility_interruption', 'NO')
        ws.cell(row=row, column=1, value="Will there be Utility interruption?").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(utility == 'YES'))
        ws.cell(row=row, column=3, value="YES")
        ws.cell(row=row, column=4, value=create_checkbox(utility == 'NO'))
        ws.cell(row=row, column=5, value="NO")
        ws.cell(row=row, column=6, value=create_checkbox(utility == 'N/A'))
        ws.cell(row=row, column=7, value="N/A")
        row += 1
        
        if utility == 'YES':
            ws.cell(row=row, column=1, value="Specify affected utilities and affected areas").font = label_font
            ws.merge_cells(f'B{row}:G{row}')
            ws.cell(row=row, column=2, value=data.get('affected_utilities', ''))
            row += 1
        
        # Waste Generation
        waste = data.get('waste_generation', 'NO')
        ws.cell(row=row, column=1, value="Will there be waste generation?").font = label_font
        ws.cell(row=row, column=2, value=create_checkbox(waste == 'YES'))
        ws.cell(row=row, column=3, value="YES")
        ws.cell(row=row, column=4, value=create_checkbox(waste == 'NO'))
        ws.cell(row=row, column=5, value="NO")
        row += 1
        
        if waste == 'YES':
            ws.cell(row=row, column=1, value="Identify and list possible waste generated").font = label_font
            ws.merge_cells(f'B{row}:G{row}')
            ws.cell(row=row, column=2, value=data.get('waste_list', 'Boxes/Packaging materials of NOKIA equipments.'))
            row += 1
        
        row += 1
        
        # ============================================
        # TOOLS AND MATERIALS
        # ============================================
        ws.cell(row=row, column=1, value="List of tools and materials to be used").font = title_font
        row += 1
        
        tools = data.get('tools_materials', ['PPE', 'FIRST AID KIT', 'ODF', 'OPM', 'OTDR', 'Patchcord', 'Fusion Splicer', 'Cleaning Kit'])
        for tool in tools:
            ws.cell(row=row, column=1, value=tool)
            row += 1
        
        row += 1
        
        # ============================================
        # JOB HAZARD ASSESSMENT
        # ============================================
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="JOB HAZARD ASSESSMENT")
        cell.font = title_font
        cell.alignment = center_alignment
        row += 1
        
        # JHA Table Header
        headers = ['Job Step', 'Hazard', '', 'Recommended Controls', '', '', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = label_font
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        # JHA Entries
        jha_steps = data.get('jha_steps', [])
        if jha_steps:
            for step in jha_steps:
                ws.cell(row=row, column=1, value=step.get('step', '')).border = thin_border
                ws.cell(row=row, column=2, value=step.get('hazard', '')).border = thin_border
                ws.cell(row=row, column=4, value=step.get('controls', '')).border = thin_border
                row += 1
        else:
            # Default JHA entries
            default_jha = [
                {'step': 'Site Access', 'hazard': 'Slip, trip, and fall from uneven surface', 'controls': 'Coordinate with lessor/UDI security prior to entry. Ensure all permits are on hand.'},
                {'step': 'Prepare Work Area', 'hazard': 'Trips/Falls: Uneven surfaces, debris', 'controls': 'Clear work area of debris, ensure adequate lighting, wear safety shoes.'},
                {'step': 'Handling Fiber Optic cables', 'hazard': 'Cuts/Lacerations: Sharp edges', 'controls': 'Wear cut-resistant gloves. Handle fibers with care.'}
            ]
            for step in default_jha:
                ws.cell(row=row, column=1, value=step['step']).border = thin_border
                ws.cell(row=row, column=2, value=step['hazard']).border = thin_border
                ws.cell(row=row, column=4, value=step['controls']).border = thin_border
                row += 1
        
        row += 1
        
        # ============================================
        # REQUIRED PPE
        # ============================================
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws.cell(row=row, column=1, value="REQUIRED PPE")
        cell.font = title_font
        cell.alignment = center_alignment
        row += 1
        
        ppe_required = data.get('ppe_required', ['Safety Shoes', 'Hardhat', 'Body Harness', 'Gloves'])
        ppe_items = ['Safety Shoes', 'Hardhat', 'Body Harness', 'Gloves', 'Welding Mask', 'N95 Masks', 'Goggles', 'Other PPE']
        
        for i, ppe in enumerate(ppe_items[:4]):
            col = get_column_letter(i + 1)
            ws.cell(row=row, column=i+1, value=create_checkbox(ppe in ppe_required))
            ws.cell(row=row+1, column=i+1, value=ppe).font = label_font
        row += 2
        
        for i, ppe in enumerate(ppe_items[4:]):
            col = get_column_letter(i + 1)
            ws.cell(row=row, column=i+1, value=create_checkbox(ppe in ppe_required))
            ws.cell(row=row+1, column=i+1, value=ppe).font = label_font
        
        if 'Other PPE' in ppe_required:
            ws.cell(row=row+2, column=5, value=data.get('other_ppe_text', ''))
        
        row += 4
        
        # ============================================
        # LIST OF WORKERS
        # ============================================
        ws.cell(row=row, column=1, value="List of workers").font = title_font
        row += 1
        
        workers = data.get('workers', [])
        for i, worker in enumerate(workers[:16]):
            col = 1 if i % 2 == 0 else 2
            if i % 2 == 0 and i > 0:
                row += 1
            ws.cell(row=row, column=col, value=worker)
        row += 2
        
        # ============================================
        # ACKNOWLEDGEMENT
        # ============================================
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="ACKNOWLEDGEMENT")
        cell.font = title_font
        cell.alignment = center_alignment
        row += 1
        
        ws.cell(row=row, column=1, value="Prepared By:").font = label_font
        ws.cell(row=row, column=2, value=data.get('prepared_by', 'RONNIE ALVIN-CHIU'))
        ws.cell(row=row, column=3, value="Noted By").font = label_font
        ws.cell(row=row, column=4, value=data.get('noted_by', 'John Carlo Rabanes'))
        ws.cell(row=row, column=5, value="Approved By").font = label_font
        row += 1
        
        ws.cell(row=row, column=1, value="Project Safety Officer")
        ws.cell(row=row, column=2, value="MNO/Project In-Charge")
        ws.cell(row=row, column=4, value=create_checkbox(data.get('approved_status') == 'YES'))
        ws.cell(row=row, column=5, value="YES")
        ws.cell(row=row, column=6, value=create_checkbox(data.get('approved_status') == 'NO'))
        ws.cell(row=row, column=7, value="NO")
        row += 1
        
        ws.cell(row=row, column=5, value=data.get('safety_officer_approval', 'PTG MIDC O&M Regional Manager'))
        row += 2
        
        # ============================================
        # FOOTER
        # ============================================
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws.cell(row=row, column=1, value="The Vendor/Contractor takes full ownership in ensuring the Health and Safety of its workers onsite by carefully assessing the activity. That the assigned Safety Officer and the Project In-Charge had identified the corresponding hazards and their appropriate safety controls. And all assigned personnel to this activity were made aware of their tasks, and the hazards and commits to carry out the safety controls. Any activity not included in the above scope will not be executed without a New Health and Safety Work Permit submitted.")
        cell.alignment = left_alignment
        cell.font = Font(name='Arial', size=9, italic=True)
        
        # Save the workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except Exception as e:
        st.error(f"Error creating Excel file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

def get_excel_download_link(file_data, filename):
    """Generate download link for Excel file"""
    b64 = base64.b64encode(file_data).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}" style="background-color: #4CAF50; color: white; padding: 12px 24px; text-decoration: none; border-radius: 4px; display: inline-block; font-weight: bold; font-size: 16px;">📥 Download Completed Excel File</a>'
    return href

def main():
    # Create two columns for layout
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📋 Project Details")
        
        sub_contractor = st.text_input("Name of Sub Contractor", value="Ultegra Supplies and Services")
        requesting_vendor = st.text_input("Requesting Vendor", value="NOKIA SHANGHAI BELL")
        project_in_charge = st.text_input("Project In-Charge", value="")
        person_in_charge = st.text_input("Person In-charge", value="John Carlo Rabanes")
        safety_officer = st.text_input("Project Safety Officer", value="RONNIE ALVIN CHIU")
        work_schedule = st.text_input("Work Schedule", value="")
        work_time_period = st.text_input("Work Time Period", value="")
        project_name = st.text_input("Project Name", value="FTTH HORIZONTAL")
        work_location = st.text_input("Work Location", value="MIN624_TS ORAPOBBUTUANAGN")
        tower_type = st.text_input("Tower Type", value="ground base")
        brief_description = st.text_area("Brief Description of Work", value="SFP link upgrade, Survey", height=68)
        
        col1a, col1b = st.columns(2)
        with col1a:
            start_date = st.date_input("Start Date", value=datetime(2026, 8, 10))
            start_time = st.text_input("Start Time", value="08:00AM")
        with col1b:
            end_date = st.date_input("End Date", value=datetime(2026, 9, 10))
            end_time = st.text_input("End Time", value="08:00PM")
        
        st.subheader("🛠️ Work Details")
        
        high_risk = st.radio("Is work to be done with high risk?", ["NO", "YES"])
        
        if high_risk == "YES":
            with st.expander("Work at Heights", expanded=True):
                col_wh1, col_wh2 = st.columns(2)
                with col_wh1:
                    work_at_heights = st.checkbox("Work at Heights")
                    scaffold = st.checkbox("Scaffold")
                    scaffold_cert = st.text_input("NCII Cert of Scaffold Erector")
                with col_wh2:
                    ladder = st.checkbox("Ladder")
                    tower = st.checkbox("Tower")
                    wah_rigger_cert = st.text_input("WAH Rigger Certificate")
                
                scaffold_components = st.checkbox("Scaffold components available")
                workers_fit = st.checkbox("Workers physically fit")
            
            with st.expander("Electrical Works", expanded=True):
                electrical_works = st.checkbox("Electrical Works")
                electrician_cert = st.text_input("NCII Cert of Electrician or ID of REE/RME")
                loto_device = st.checkbox("LOTO Device")
                insulated_tools = st.checkbox("Insulated Tools")
            
            with st.expander("Heavy Lifting / Excavation", expanded=True):
                operator_cert = st.text_input("NCII Cert of Operator")
                rigger_cert = st.text_input("Cert of Lift Rigger")
                heavy_eqpt_cert = st.text_input("3rd Party Certification of Heavy Eqpt")
            
            with st.expander("Confined Space Works", expanded=True):
                confined_space = st.checkbox("Confined Space Works")
                col_cs1, col_cs2 = st.columns(2)
                with col_cs1:
                    scba_cert = st.text_input("Certificate of SCBA Operator")
                    ventilation_eqpt = st.text_input("Ventilation Equipment")
                    flash_arrester = st.checkbox("OxyFuel flash back arrester installed")
                with col_cs2:
                    fire_blanket = st.checkbox("Fire Blanket")
                    o2_detector = st.text_input("O2 and Gas Detector")
                    safety_line = st.text_input("Safety Line")
            
            with st.expander("Harmful Substances", expanded=True):
                harmful_substance = st.radio("Is there any harmful substance or nuisance release?", ["NO", "YES"])
                if harmful_substance == "YES":
                    col_h1, col_h2 = st.columns(2)
                    with col_h1:
                        fumes = st.checkbox("Fumes")
                        dust = st.checkbox("Dust")
                        sparks = st.checkbox("Sparks")
                    with col_h2:
                        odors = st.checkbox("Offensive Odors")
                        noise = st.checkbox("Noise")
                        other_harmful = st.text_input("Others:")
            
            with st.expander("Utility Interruption", expanded=True):
                utility_interruption = st.radio("Will there be Utility interruption?", ["NO", "YES", "N/A"])
                if utility_interruption == "YES":
                    affected_utilities = st.text_area("Specify affected utilities and affected areas")
        else:
            work_at_heights = scaffold = ladder = tower = False
            scaffold_cert = wah_rigger_cert = ""
            scaffold_components = workers_fit = False
            electrical_works = False
            electrician_cert = loto_device = insulated_tools = ""
            operator_cert = rigger_cert = heavy_eqpt_cert = ""
            confined_space = False
            scba_cert = ventilation_eqpt = flash_arrester = fire_blanket = ""
            o2_detector = safety_line = ""
            harmful_substance = "NO"
            fumes = odors = dust = noise = sparks = False
            other_harmful = ""
            utility_interruption = "NO"
            affected_utilities = ""
        
        # Waste Generation
        st.subheader("♻️ Waste Generation")
        waste_generation = st.radio("Will there be waste generation?", ["NO", "YES"])
        waste_list = ""
        if waste_generation == "YES":
            waste_list = st.text_area("Identify and list possible waste generated", 
                                     value="Boxes/Packaging materials of NOKIA equipments.", height=80)
        
        st.subheader("📌 JOB HAZARD ASSESSMENT (Bottom Table)")
        
        num_jha = st.number_input("Number of JHA entries", min_value=0, max_value=10, value=0)
        
        jha_steps = []
        for i in range(num_jha):
            with st.expander(f"JHA Entry {i+1}", expanded=i==0):
                step = st.text_input(f"Job Step {i+1}", key=f"jha_step_{i}")
                hazard = st.text_input(f"Hazard {i+1}", key=f"jha_hazard_{i}")
                controls = st.text_area(f"Controls {i+1}", key=f"jha_controls_{i}", height=60)
                if step and hazard and controls:
                    jha_steps.append({
                        'step': step,
                        'hazard': hazard,
                        'controls': controls
                    })
    
    with col2:
        st.subheader("🔧 Tools and Materials")
        tools_text = st.text_area("List tools and materials (one per line)", 
                                  value="PPE\nFIRST AID KIT\nODF\nOPM\nOTDR\nPatchcord\nFusion Splicer\nCleaning Kit", height=150)
        tools = [t.strip() for t in tools_text.split('\n') if t.strip()]
        
        st.subheader("🦺 Required PPE")
        ppe_options = ['Safety Shoes', 'Hardhat', 'Body Harness', 'Gloves', 'Welding Mask', 'N95 Masks', 'Goggles', 'Other PPE']
        ppe_required = st.multiselect("Select required PPE", ppe_options, default=['Safety Shoes', 'Hardhat', 'Body Harness', 'Gloves'])
        other_ppe_text = ""
        if 'Other PPE' in ppe_required:
            other_ppe_text = st.text_input("Other PPE details")
        
        st.subheader("👷 List of Workers")
        workers_text = st.text_area("List workers (one per line)",
                                   value="CABRAL, RAYMOND R.\nCANETE, GREMAR L.\nORBITA, RICHARD M.\nPELISCO, KENNETH KIM L.\nPRIETO, JASON D.\nPRIETO, KEITH JHIRVINE D.\nPRIETO, SEAN DALE D.\nSABAD, VIRGILIO JR M.\nSUMILE, JUSTINE KIM J.\nTRANA, AQUILINO B.\nTRANA, TRANQUILINO B.\nRABANES, JOHN CARLO", height=200)
        workers = [w.strip() for w in workers_text.split('\n') if w.strip()]
        
        st.subheader("✅ Acknowledgement")
        prepared_by = st.text_input("Prepared By", value="RONNIE ALVIN-CHIU")
        noted_by = st.text_input("Endorsed By", value="John Carlo Rabanes")
        approved_by = st.radio("Approved By", ["YES", "NO"])
        safety_officer_approval = st.text_input("MIDC O&M Regional Manager", value="PTG MIDC O&M Regional Manager")

    # Collect all data
    data = {
        'sub_contractor': sub_contractor,
        'requesting_vendor': requesting_vendor,
        'project_in_charge': project_in_charge,
        'person_in_charge': person_in_charge,
        'safety_officer': safety_officer,
        'work_schedule': work_schedule,
        'work_time_period': work_time_period,
        'project_name': project_name,
        'work_location': work_location,
        'tower_type': tower_type,
        'start_date': start_date.strftime('%m/%d/%Y'),
        'end_date': end_date.strftime('%m/%d/%Y'),
        'start_time': start_time,
        'end_time': end_time,
        'brief_description': brief_description,
        'high_risk': high_risk,
        'work_at_heights': locals().get('work_at_heights', False),
        'scaffold': locals().get('scaffold', False),
        'ladder': locals().get('ladder', False),
        'tower': locals().get('tower', False),
        'scaffold_cert': locals().get('scaffold_cert', ''),
        'wah_rigger_cert': locals().get('wah_rigger_cert', ''),
        'scaffold_components': locals().get('scaffold_components', False),
        'workers_fit': locals().get('workers_fit', False),
        'electrical_works': locals().get('electrical_works', False),
        'electrician_cert': locals().get('electrician_cert', ''),
        'loto_device': locals().get('loto_device', False),
        'insulated_tools': locals().get('insulated_tools', False),
        'operator_cert': locals().get('operator_cert', ''),
        'rigger_cert': locals().get('rigger_cert', ''),
        'heavy_eqpt_cert': locals().get('heavy_eqpt_cert', ''),
        'confined_space': locals().get('confined_space', False),
        'scba_cert': locals().get('scba_cert', ''),
        'ventilation_eqpt': locals().get('ventilation_eqpt', ''),
        'flash_arrester': locals().get('flash_arrester', False),
        'fire_blanket': locals().get('fire_blanket', False),
        'o2_detector': locals().get('o2_detector', ''),
        'safety_line': locals().get('safety_line', ''),
        'harmful_substance': locals().get('harmful_substance', 'NO'),
        'fumes': locals().get('fumes', False),
        'odors': locals().get('odors', False),
        'dust': locals().get('dust', False),
        'noise': locals().get('noise', False),
        'sparks': locals().get('sparks', False),
        'other_harmful': locals().get('other_harmful', ''),
        'utility_interruption': locals().get('utility_interruption', 'NO'),
        'affected_utilities': locals().get('affected_utilities', ''),
        'waste_generation': waste_generation,
        'waste_list': waste_list,
        'tools_materials': tools,
        'jha_steps': jha_steps,
        'ppe_required': ppe_required,
        'other_ppe_text': other_ppe_text,
        'workers': workers,
        'prepared_by': prepared_by,
        'noted_by': noted_by,
        'approved_status': approved_by,
        'safety_officer_approval': safety_officer_approval
    }
    
    st.markdown("---")
    
    st.info("""
    ℹ️ **This creates a clean, well-formatted Excel file from scratch with:**
    - ☑/☐ Checkboxes (Unicode characters)
    - All your data properly organized
    - Professional formatting
    - No template corruption issues
    """)
    
    # Generate Excel button
    if st.button("📥 Generate Excel File", type="primary"):
        with st.spinner("Creating Excel file..."):
            try:
                file_data = create_excel_file(data)
                if file_data:
                    filename = f"HSWP_{project_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    download_link = get_excel_download_link(file_data.getvalue(), filename)
                    
                    st.success("✅ Excel file created successfully!")
                    st.markdown(download_link, unsafe_allow_html=True)
                    
                    with st.expander("📊 Preview Data Summary"):
                        summary_data = {
                            'Project': project_name,
                            'Location': work_location,
                            'Sub Contractor': sub_contractor,
                            'Safety Officer': safety_officer,
                            'High Risk': high_risk,
                            'JHA Entries': len(jha_steps),
                            'Workers': len(workers),
                            'Tools': len(tools),
                            'PPE Items': len(ppe_required)
                        }
                        st.json(summary_data)
                else:
                    st.error("❌ Failed to create Excel file.")
                    
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                st.exception(e)
    
    # Instructions
    with st.sidebar:
        st.header("📝 Instructions")
        st.markdown("""
        1. Fill in all fields
        2. Add JHA entries as needed
        3. Select PPE
        4. List tools and workers
        5. Click Generate
        6. Download the file
        """)
        
        st.header("✅ Features")
        st.markdown("""
        - Clean, professional formatting
        - ☑/☐ checkbox style
        - All data organized
        - No template issues
        - Works reliably
        """)

if __name__ == "__main__":
    main()
